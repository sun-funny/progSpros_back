import os
import pandas as pd
import json
from datetime import datetime, date
from decimal import Decimal
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from typing import List, Tuple, Dict, Any

class CustomJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        elif isinstance(obj, Decimal):
            return float(obj)
        # Add other non-serializable types here
        return super().default(obj)

""" GET_DATA """
def get_data_exl(query, shown_columns, yearfrom: int, yearto: int, otrasl_total: str, sum_pr: float = None) -> List[Dict[str, Any]]:
    """Основная функция обработки данных для формирования структуры отчета"""
    # Конфигурация стилей
    STYLE_CONFIG = {
        'header': {
            'действующие потребители': 's01',
            'перспективные потребители': 's03',
            'потенциальные потребители': 's05'
        },
        'subheader': {
            'действующие потребители': 's011',
            'перспективные потребители': 's031',
            'потенциальные потребители': 's051'
        },
        'special': 's02',
        'region_header': {
            'действующие потребители': 's04',
            'ожидаемые перспективные потребители':'s04',
            'потенциальные перспективные потребители': 's05'
        }
    }
    # Инициализация данных
    years = [str(y) for y in range(yearfrom, yearto + 1)]
    agg_dict = {f'y{y}': 'sum' for y in years}
    agg_dict['prirost'] = 'sum'

    others_columns = shown_columns + ['otrasl_ord'] if 'otrasl_ord' not in shown_columns else shown_columns

    try:
        results = query.session.execute(query).fetchall()
        df = pd.DataFrame(results, columns=[col['name'] for col in query.column_descriptions])
        if sum_pr is not None:
            #df = df[df['prirost'] > sum_pr]
            df = df[df[f'y{yearto}'] > sum_pr]

    except Exception as e:
        print(f"Error reading SQL data: {e}")
        raise
    processed_df = preprocess_data(df, years, others_columns)
    result = []

    # Обработка основных групп
    process_main_groups(processed_df, years, others_columns, agg_dict, STYLE_CONFIG, result, otrasl_total)

    # Обработка программы газификации для всех данных (один раз в конце основных групп)
    process_gaz_program(processed_df, years, others_columns, agg_dict, 's06', result, region_name=None)

    # Обработка специальных категорий для всех данных (один раз в конце основных групп)
    process_special_category(processed_df, years, 'expect', 'ожидаемый', agg_dict, STYLE_CONFIG['special'], result)
    process_special_category(processed_df, years, 'maximum', 'максимальный', agg_dict, STYLE_CONFIG['special'], result)

    # Обработка региональных данных
    process_regional_data(processed_df, years, others_columns, agg_dict, STYLE_CONFIG, result, otrasl_total)

    return result

def process_gaz_program(df: pd.DataFrame, years: List[str], others_columns: List[str],
                        agg_dict: Dict[str, str], style: str, result: List,
                        region_name: str = None) -> None:
    """Обработка программы газификации"""
    gaz_group = df[df['infr'] == 'V']
    if not gaz_group.empty:
        gaz_sum = gaz_group.agg({**{f'y{y}': 'sum' for y in years}, 'prirost': 'sum'})

        # Заголовок в зависимости от региона
        if region_name:
            title = f"{region_name} по Программе газификации"
        else:
            title = "ВСЕГО по Программе газификации"

        result.append(
            create_header_row_with_sum(title, gaz_sum, others_columns, years, style)
        )

def preprocess_data(df: pd.DataFrame, years: List[str], others_columns: List[str]) -> pd.DataFrame:
    """Предварительная обработка данных"""
    # Сначала очищаем данные для специальных контрагентов
    mask = df['contragent'].isin(['Действующие потребители', 'Прочие потребители'])
    df.loc[mask, others_columns] = ''

    # Затем преобразуем категории
    category_mappings = {
        'ver_real_level1': {
            '0 - действующие потребители': 'действующие потребители',
            '1 - высокая вероятность реализации проекта': 'ожидаемые перспективные потребители',
            '2 - средняя вероятность реализации проекта': 'ожидаемые перспективные потребители',
            '3 - низкая вероятность реализации проекта': 'потенциальные перспективные потребители',
        },
        'ver_real_level2': {
            'действующие потребители': 'действующие потребители',
            'ожидаемые перспективные потребители': 'перспективные потребители',
            'потенциальные перспективные потребители': 'потенциальные потребители',
        },
        'expect': {
            'действующие потребители': 'ожидаемый',
            'ожидаемые перспективные потребители': 'ожидаемый'
        },
        'maximum': {
            'действующие потребители': 'максимальный',
            'ожидаемые перспективные потребители': 'максимальный',
            'потенциальные перспективные потребители': 'максимальный'
        }
    }

    columns_mappings = {
        'ver_real_level1': 'ver_real',
        'ver_real_level2': 'ver_real_level1',
        'expect': 'ver_real_level1',
        'maximum': 'ver_real_level1'
    }

    for col, mapping in category_mappings.items():
        df[col] = df[columns_mappings[col]].replace(mapping).fillna('Ошибка')

    return df


def process_main_groups(df: pd.DataFrame, years: List[str], others_columns: List[str],
                        agg_dict: Dict[str, str], styles: Dict, result: List, otrasl_total: str) -> None:
    # 1. Действующие потребители (s01 + s011)
    active_group = df[df['ver_real_level2'] == "действующие потребители"]
    if not active_group.empty:
        active_sum = active_group.groupby(['ver_real_level2']).agg(agg_dict).reset_index().iloc[0]
        result.append(create_header_row_with_sum(
            "ВСЕГО, действующие потребители",
            active_sum, others_columns, years,
            styles['header']["действующие потребители"]
        ))
        process_otrasl_total_row(active_group, years, agg_dict, 's011', result, otrasl_total)

    # 2. Перспективные + Потенциальные вместе как s03
    pers_group = df[df['ver_real_level2'].isin(["перспективные потребители", "потенциальные потребители"])]
    if not pers_group.empty:
        pers_sum = pers_group.agg({**{f'y{y}': 'sum' for y in years}, 'prirost': 'sum'})
        result.append(create_header_row_with_sum(
            "ВСЕГО, перспективные потребители, в т.ч.:",
            pers_sum, others_columns, years, 's03'
        ))

        # Подгруппы внутри → ожидаемые (s031) и потенциальные (s051)
        for subgroup_name, subgroup_data in pers_group.groupby('ver_real_level2'):
            if subgroup_name == "перспективные потребители":
                display_name, style = "ожидаемые потребители", "s031"
            elif subgroup_name == "потенциальные потребители":
                display_name, style = "потенциальные потребители", "s051"
            else:
                display_name, style = subgroup_name, "s05"  # fallback

            subgroup_sum = subgroup_data.agg({**{f'y{y}': 'sum' for y in years}, 'prirost': 'sum'})
            result.append(create_data_row(display_name, subgroup_sum, years, style))

            # внутри них → в.т.ч. Население (s011)
            process_otrasl_total_row(subgroup_data, years, agg_dict, 's011', result, otrasl_total)

def process_otrasl_total_row(group: pd.DataFrame, years: List[str], agg_dict: Dict[str, str],
                             style: str, result: List, otrasl_total: str) -> None:
    """
    Добавляет строку "в.т.ч. otrasl_total" для подгруппы.
    """
    if otrasl_total in group['otrasl'].unique():
        otrasl_group = group[group['otrasl'] == otrasl_total]
        otrasl_sum = otrasl_group.agg({**{f'y{y}': 'sum' for y in years}, 'prirost': 'sum'})

        result.append(create_data_row(f"   в т.ч. {otrasl_total}", otrasl_sum, years, style))

def create_header_row_with_sum(title: str, sum_data: Any, others_columns: List[str], years: List[str], style: str) -> Dict:
    """Создает строку-заголовок с суммами"""
    return {
        'contragent': f'{title}',
        **{col: '' for col in others_columns},
        **{f'y{y}': safe_float(safe_getattr(sum_data, f'y{y}', 0)) for y in years},
        'prirost': safe_float(safe_getattr(sum_data, 'prirost', 0)),
        'style': style
    }

def process_regional_data(df: pd.DataFrame, years: List[str], others_columns: List[str],
                          agg_dict: Dict[str, str], styles: Dict, result: List, otrasl_total: str) -> None:
    """Обработка региональных данных"""
    for region, region_group in df.groupby('region'):

        # 1. ДЕЙСТВУЮЩИЕ (s01 → s011)
        active_group = region_group[region_group['ver_real_level2'] == "действующие потребители"]
        if not active_group.empty:
            active_sum = active_group.agg({**{f'y{y}': 'sum' for y in years}, 'prirost': 'sum'})
            result.append(create_header_row_with_sum(f"{region}, действующие потребители",
                                                     active_sum, others_columns, years, 's01'))
            process_otrasl_total_row(active_group, years, agg_dict, 's011', result, otrasl_total)
           #process_subgroups(active_group, years, agg_dict, 's011', result, otrasl_total)

        # 2. ПЕРСПЕКТИВНЫЕ + ПОТЕНЦИАЛЬНЫЕ (s03 = s031 + s051 → s011)
        pers_group = region_group[region_group['ver_real_level2'].isin(["перспективные потребители",
                                                                       "потенциальные потребители"])]
        if not pers_group.empty:
            pers_sum = pers_group.agg({**{f'y{y}': 'sum' for y in years}, 'prirost': 'sum'})
            result.append(create_header_row_with_sum(f"{region}, перспективные потребители, в т.ч.:",
                                                     pers_sum, others_columns, years, 's03'))

            for subgroup_name, subgroup_data in pers_group.groupby('ver_real_level2'):

                if subgroup_name == "перспективные потребители":
                    display_name, style = "ожидаемые потребители", "s031"
                elif subgroup_name == "потенциальные потребители":
                    display_name, style = "потенциальные потребители", "s051"
                else:
                    display_name, style = subgroup_name, "s05"

                subgroup_sum = subgroup_data.agg({**{f'y{y}': 'sum' for y in years}, 'prirost': 'sum'})

                # строка s031 / s051
                result.append(create_data_row(display_name, subgroup_sum, years, style))

                # внутри них → s011
                process_otrasl_total_row(subgroup_data, years, agg_dict, 's011', result, otrasl_total)
                #process_subgroups(subgroup_data, years, agg_dict, 's011', result, otrasl_total)

        # 3. ПРОГРАММА ГАЗИФИКАЦИИ (s06)
        process_gaz_program(region_group, years, others_columns, agg_dict, 's06', result, region_name=region)

        # 4. ОЖИДАЕМЫЙ (s02)
        process_special_category(region_group, years, 'expect', 'ожидаемый',
                                 agg_dict, styles['special'], result)

        # 5. МАКСИМАЛЬНЫЙ (s02)
        process_special_category(region_group, years, 'maximum', 'максимальный',
                                 agg_dict, styles['special'], result)

        # 6–8. Итоговые блоки по ver_real_level2 (s04 → s041)
        process_ver_real_level2_groups(region_group, years, others_columns,
                                       agg_dict, styles, result, otrasl_total)

def sort_subgroup(df: pd.DataFrame) -> pd.DataFrame:
    """Сортировка подгрупп: сначала по otrasl_ord (пустые в конце), затем по контрагенту"""
    return (
        df.assign(
            _otrasl_sort=lambda x: x['otrasl_ord'].apply(
                lambda val: (1 if pd.isna(val) or val == '' else 0, val if not pd.isna(val) and val != '' else 0)
            )
        )
        .sort_values(
            by=['_otrasl_sort', 'otrasl_ord', 'contragent'],
            ascending=[True, True, True]  # Сначала непустые отрасли, затем пустые
        )
        .drop(columns=['_otrasl_sort'])
    )
def process_ver_real_level2_groups(df: pd.DataFrame, years: List[str], others_columns: List[str],
                                   agg_dict: Dict[str, str], styles: Dict, result: List, otrasl_total: str) -> None:
    """Обработка групп по ver_real_level2 с выводом сумм"""
    for level_name, group in df.groupby('ver_real_level2'):
        # Приводим названия для стиля s04
        display_name = level_name
        if level_name == "действующие потребители":
            display_name = "Действующие потребители"
        elif level_name == "перспективные потребители":
            display_name = "Ожидаемые потребители"
        elif level_name == "потенциальные потребители":
            display_name = "Потенциальные потребители"
        # Вычисляем суммы для группы
        group_sum = group.groupby(['ver_real_level2']).agg(agg_dict).reset_index().iloc[0]

        # Добавляем заголовок с суммами с стилем s04 для ver_real_level2
        result.append({
            'contragent': display_name,
            **{col: '' for col in others_columns},
            **{f'y{y}': safe_float(safe_getattr(group_sum, f'y{y}', 0)) for y in years},
            'prirost': safe_float(safe_getattr(group_sum, 'prirost', 0)),
            'style': 's04'  # Стиль s04 для ver_real_level2
        })

        # Добавляем детализированные строки
        subgroup = (
            group.groupby(['contragent'] + others_columns)
            .agg(agg_dict)
            .reset_index()
        )
        subgroup = sort_subgroup(subgroup)

        for row in subgroup.itertuples():
            result.append(create_detailed_row(row, years, others_columns))

def safe_getattr(obj, attr, default=0):
    """Безопасное получение атрибута объекта"""
    try:
        return getattr(obj, attr, default)
    except (AttributeError, IndexError):
        return default

def process_ver_real_level1_groups(df: pd.DataFrame, years: List[str], others_columns: List[str],
                                   agg_dict: Dict[str, str], styles: Dict, result: List, otrasl_total: str) -> None:
    """Обработка групп по ver_real_level1 с выводом сумм"""
    for level_name, group in df.groupby('ver_real_level1'):
        # Вычисляем суммы для группы
        group_sum = group.groupby(['ver_real_level1']).agg(agg_dict).reset_index().iloc[0]

        # Добавляем заголовок с суммами с стилем s031 для ver_real_level1
        result.append({
            'contragent': level_name,
            **{col: '' for col in others_columns},
            **{f'y{y}': safe_float(safe_getattr(group_sum, f'y{y}', 0)) for y in years},
            'prirost': safe_float(safe_getattr(group_sum, 'prirost', 0)),
            'style': 's031'  # Стиль s031 для ver_real_level1
        })

        # Добавляем детализированные строки
        subgroup = (
            group.groupby(['contragent'] + others_columns + ['otrasl_ord'])
            .agg(agg_dict)
            .reset_index()
        )
        subgroup = sort_subgroup(subgroup)

        for row in subgroup.itertuples():
            result.append(create_detailed_row(row, years, others_columns))


def process_subgroups(group: pd.DataFrame, years: List[str], agg_dict: Dict[str, str],
                      style: str, result: List, otrasl_total: str) -> None:
    for subgroup_name, subgroup_data in group.groupby('ver_real_level1'):
        subgroup_sum = subgroup_data.agg({**{f'y{y}': 'sum' for y in years}, 'prirost': 'sum'})

        # подгруппа (уровень 2) → стиль s031
        result.append(create_data_row(subgroup_name, subgroup_sum, years, style))

        # в.т.ч. otrasl_total (уровень 3) → стиль s011
        process_otrasl_total_row(subgroup_data, years, agg_dict, 's011', result, otrasl_total)

def process_special_category(df: pd.DataFrame, years: List[str], category: str, filter: str,
                             agg_dict: Dict[str, str], style: str, result: List) -> None:
    """Обработка специальных категорий"""
    filtered = df[df[category] == filter]
    grouped = filtered.groupby([category]).agg(agg_dict).reset_index()

    name_map = {
        "ожидаемый": "Ожидаемый сценарий спроса (действ.+ожидаемые потребители)",
        "максимальный": "Максимальный сценарий спроса (действ.+ожидаемые+потенц.потребители)"
    }

    for row in grouped.itertuples():
        title = name_map.get(getattr(row, category), getattr(row, category))
        result.append(create_data_row(title, row, years, style))

def create_header_row(title: str, others_columns: List[str], years: List[str], style: str) -> Dict:
    """Создает строку-заголовок"""
    return {
        'contragent': f'{title}',
        **{col: '' for col in others_columns},
        **{f'y{y}': '' for y in years},
        'prirost': '',
        'style': style
    }

def create_data_row(title: str, row_data: Any, years: List[str], style: str) -> Dict:
    """Создает строку с данными"""
    return {
        'contragent': title,
        **{f'y{y}': safe_float(getattr(row_data, f'y{y}')) for y in years},
        'prirost': safe_float(getattr(row_data, 'prirost', 0)),
        'style': style
    }

def create_detailed_row(row_data: Any, years: List[str], others_columns: List[str]) -> Dict:
    """Создает детализированную строку с данными"""
    return {
        'contragent': row_data.contragent,
        **{col: getattr(row_data, col, None) for col in others_columns},
        'otrasl_ord': getattr(row_data, 'otrasl_ord', None),
        **{f'y{y}': safe_float(getattr(row_data, f'y{y}')) for y in years},
        'prirost': row_data.prirost,
        'style': 's041'
    }

def safe_float(value: Any) -> float:
    return float(value) if value is not None else 0.0

""" GET_EXCEL """
def get_excel(result, yearfrom, yearto, fo_params=None):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
    from openpyxl.utils import get_column_letter
    import json
    import pandas as pd

    STYLES = {
        "s01": {
            "fill": PatternFill("solid", fgColor="FFC000"),
            "font": Font(name='Times New Roman', size=22, bold=True),
            "alignment": Alignment(horizontal='center', vertical='center', wrap_text=True)
        },
        "contragent_s01": {
            "fill": PatternFill("solid", fgColor="FFC000"),
            "font": Font(name='Times New Roman', size=22, bold=True, color='000000'),
            "alignment": Alignment(indent=1, vertical='center', wrap_text=True)
        },
        "s011": {
            "fill": PatternFill("solid", fgColor="FFFFFF"),
            "font": Font(name='Times New Roman', size=18, italic=True),
            "alignment": Alignment(horizontal='center', vertical='center', wrap_text=True)
        },
        "contragent_s011": {
            "fill": PatternFill("solid", fgColor="FFFFFF"),
            "font": Font(name='Times New Roman', size=18, italic=True, color='000000'),
            "alignment": Alignment(indent=2, vertical='center', wrap_text=True)
        },
        "s03": {
            "fill": PatternFill("solid", fgColor="9BC2E6"),
            "font": Font(name='Times New Roman', size=22, bold=True),
            "alignment": Alignment(horizontal='center', vertical='center', wrap_text=True)
        },
        "contragent_s03": {
            "fill": PatternFill("solid", fgColor="9BC2E6"),
            "font": Font(name='Times New Roman', size=22, bold=True, color='000000'),
            "alignment": Alignment(indent=1, vertical='center', wrap_text=True)
        },
        "s031": {
            "fill": PatternFill("solid", fgColor="FFFFFF"),
            "font": Font(name='Times New Roman', size=22, bold=True),
            "alignment": Alignment(horizontal='center', vertical='center', wrap_text=True)
        },
        "contragent_s031": {
            "fill": PatternFill("solid", fgColor="FFFFFF"),
            "font": Font(name='Times New Roman', size=22, bold=True, color='000000'),
            "alignment": Alignment(indent=2, vertical='center', wrap_text=True)
        },
        "s051": {
            "fill": PatternFill("solid", fgColor="FFFFFF"),
            "font": Font(name='Times New Roman', size=22, bold=True),
            "alignment": Alignment(horizontal='center', vertical='center', wrap_text=True)
        },
        "contragent_s051": {
            "fill": PatternFill("solid", fgColor="FFFFFF"),
            "font": Font(name='Times New Roman', size=22, bold=True, color='000000'),
            "alignment": Alignment(indent=2, vertical='center', wrap_text=True)
        },
        "s02": {
            "fill": PatternFill("solid", fgColor="FFFBB1"),
            "font": Font(name='Times New Roman', size=20),
            "alignment": Alignment(horizontal='center', vertical='center', wrap_text=True)
        },
        "contragent_s02": {
            "fill": PatternFill("solid", fgColor="FFFBB1"),
            "font": Font(name='Times New Roman', size=20, bold=True, color='000000'),
            "alignment": Alignment(indent=0, vertical='center', wrap_text=True)
        },
        "s04": {
            "fill": PatternFill("solid", fgColor="FFDAB9"),
            "font": Font(name='Times New Roman', size=20, bold=True),
            "alignment": Alignment(horizontal='center', vertical='center', wrap_text=True)
        },
        "contragent_s04": {
            "fill": PatternFill("solid", fgColor="FFDAB9"),
            "font": Font(name='Times New Roman', size=20, bold=True, color='000000'),
            "alignment": Alignment(vertical='center', wrap_text=True)
        },
        "s041": {
            "fill": PatternFill("solid", fgColor="FFFFFF"),
            "font": Font(name='Times New Roman', size=20),
            "alignment": Alignment(horizontal='center', vertical='center', wrap_text=True)
        },
        "contragent_s041": {
            "fill": PatternFill("solid", fgColor="FFFFFF"),
            "font": Font(name='Times New Roman', size=20, color='000000'),
            "alignment": Alignment(indent=1, vertical='center', wrap_text=True)
        },
        "s05": {
            "fill": PatternFill("solid", fgColor="EDEDED"),
            "font": Font(name='Times New Roman', size=22, bold=True),
            "alignment": Alignment(horizontal='center', vertical='center', wrap_text=True)
        },
        "contragent_s05": {
            "fill": PatternFill("solid", fgColor="EDEDED"),
            "font": Font(name='Times New Roman', size=22, bold=True, italic=True, color='000000'),
            "alignment": Alignment(indent=0, vertical='center', wrap_text=True)
        },
        "all": {
            "font": Font(name='Times New Roman', size=11),
            "alignment": Alignment(horizontal='center', vertical='center', wrap_text=True)
        },
        "year": {
            "font": Font(name='Times New Roman', size=11, color='000000'),
            "alignment": Alignment(vertical='center', wrap_text=True)
        },
        "s06": {
            "fill": PatternFill("solid", fgColor="E2EFD8"),
            "font": Font(name='Times New Roman', size=24, bold=False),
            "alignment": Alignment(horizontal='center', vertical='center', wrap_text=True)
        },
        "contragent_s06": {
            "fill": PatternFill("solid", fgColor="E2EFD8"),
            "font": Font(name='Times New Roman', size=24, bold=False, color='000000'),
            "alignment": Alignment(indent=1, vertical='center', wrap_text=True)
        },
    }

    RULES = [
        # Правила для столбца contragent
        {"columns": ["contragent"], "condition": lambda v: v == "s01", "style": "contragent_s01"},
        {"columns": ["contragent"], "condition": lambda v: v == "s02", "style": "contragent_s02"},
        {"columns": ["contragent"], "condition": lambda v: v == "s03", "style": "contragent_s03"},
        {"columns": ["contragent"], "condition": lambda v: v == "s04", "style": "contragent_s04"},
        {"columns": ["contragent"], "condition": lambda v: v == "s05", "style": "contragent_s05"},
        {"columns": ["contragent"], "condition": lambda v: v == "s011", "style": "contragent_s011"},
        {"columns": ["contragent"], "condition": lambda v: v == "s031", "style": "contragent_s031"},
        {"columns": ["contragent"], "condition": lambda v: v == "s051", "style": "contragent_s051"},
        {"columns": ["contragent"], "condition": lambda v: v == "s041", "style": "contragent_s041"},
        {"columns": ["contragent"], "condition": lambda v: v == "s06", "style": "contragent_s06"},

        # Правила для числовых ячеек (годы и прирост)
        {"columns": [f"y{y}" for y in range(yearfrom, yearto + 1)] + ["prirost"], "condition": lambda v: v == "s01",
         "style": "s01"},
        {"columns": [f"y{y}" for y in range(yearfrom, yearto + 1)] + ["prirost"], "condition": lambda v: v == "s011",
         "style": "s011"},
        {"columns": [f"y{y}" for y in range(yearfrom, yearto + 1)] + ["prirost"], "condition": lambda v: v == "s03",
         "style": "s03"},
        {"columns": [f"y{y}" for y in range(yearfrom, yearto + 1)] + ["prirost"], "condition": lambda v: v == "s031",
         "style": "s031"},
        {"columns": [f"y{y}" for y in range(yearfrom, yearto + 1)] + ["prirost"], "condition": lambda v: v == "s051",
         "style": "s051"},
        {"columns": [f"y{y}" for y in range(yearfrom, yearto + 1)] + ["prirost"], "condition": lambda v: v == "s02",
         "style": "s02"},
        {"columns": [f"y{y}" for y in range(yearfrom, yearto + 1)] + ["prirost"], "condition": lambda v: v == "s04",
         "style": "s04"},
        {"columns": [f"y{y}" for y in range(yearfrom, yearto + 1)] + ["prirost"], "condition": lambda v: v == "s041",
         "style": "s041"},
        {"columns": [f"y{y}" for y in range(yearfrom, yearto + 1)] + ["prirost"], "condition": lambda v: v == "s05",
         "style": "s05"},
        {"columns": [f"y{y}" for y in range(yearfrom, yearto + 1)] + ["prirost"], "condition": lambda v: v == "s06",
         "style": "s06"},

        # Правила для всех остальных столбцов (нечисловых)
        {"columns": ["*"], "condition": lambda v: v == "s01", "style": "s01",
         "exclude": ["contragent"] + [f"y{y}" for y in range(yearfrom, yearto + 1)] + ["prirost"]},
        {"columns": ["*"], "condition": lambda v: v == "s011", "style": "s011",
         "exclude": ["contragent"] + [f"y{y}" for y in range(yearfrom, yearto + 1)] + ["prirost"]},
        {"columns": ["*"], "condition": lambda v: v == "s03", "style": "s03",
         "exclude": ["contragent"] + [f"y{y}" for y in range(yearfrom, yearto + 1)] + ["prirost"]},
        {"columns": ["*"], "condition": lambda v: v == "s031", "style": "s031",
         "exclude": ["contragent"] + [f"y{y}" for y in range(yearfrom, yearto + 1)] + ["prirost"]},
        {"columns": ["*"], "condition": lambda v: v == "s051", "style": "s051",
         "exclude": ["contragent"] + [f"y{y}" for y in range(yearfrom, yearto + 1)] + ["prirost"]},
        {"columns": ["*"], "condition": lambda v: v == "s02", "style": "s02",
         "exclude": ["contragent"] + [f"y{y}" for y in range(yearfrom, yearto + 1)] + ["prirost"]},
        {"columns": ["*"], "condition": lambda v: v == "s04", "style": "s04",
         "exclude": ["contragent"] + [f"y{y}" for y in range(yearfrom, yearto + 1)] + ["prirost"]},
        {"columns": ["*"], "condition": lambda v: v == "s041", "style": "s041",
         "exclude": ["contragent"] + [f"y{y}" for y in range(yearfrom, yearto + 1)] + ["prirost"]},
        {"columns": ["*"], "condition": lambda v: v == "s05", "style": "s05",
         "exclude": ["contragent"] + [f"y{y}" for y in range(yearfrom, yearto + 1)] + ["prirost"]},
        {"columns": ["*"], "condition": lambda v: v == "s06", "style": "s06",
         "exclude": ["contragent"] + [f"y{y}" for y in range(yearfrom, yearto + 1)] + ["prirost"]},
    ]

    # Имя файла
    name = 'output.xlsx'
    # basic_path = os.path.expanduser("~/work/backend/progSpros_back/")
    #basic_path = '/opt/foresight/progSpros_back/'
    basic_path = 'C:\\Users\\user\\PycharmProjects\\backend\\progSpros_back\\'

    # Сохраняем DataFrame в Excel
    y = json.dumps(result, cls=CustomJSONEncoder)
    df = pd.read_json(y)
    df.to_excel(basic_path + name, index=False)

    wb = load_workbook(basic_path + name)
    ws = wb.active

    # Вставляем заголовок
    ws.insert_rows(1, 3)

    # Обновленные стили заголовков
    title_font = Font(name='Times New Roman', size=20, bold=True)
    subtitle_font = Font(name='Times New Roman', size=16, bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    no_border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))

    regions_text = ", ".join(fo_params) if fo_params else "Все Федеральные округа"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    ws.cell(row=1, column=1, value="Оценка прогнозного спроса на газ потребителей").font = title_font
    ws.cell(row=1, column=1).alignment = center_align

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ws.max_column)
    ws.cell(row=2, column=1, value=regions_text).font = subtitle_font
    ws.cell(row=2, column=1).alignment = center_align

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ws.max_column)

    # Убираем бордеры у первых 3 строк и выравниваем
    for r in range(1, 4):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).border = no_border
            ws.cell(row=r, column=c).alignment = center_align

    # Автоподбор высоты для заголовочных строк
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 15

    # Получаем колонки
    headers = {cell.value: idx for idx, cell in enumerate(ws[4], start=1)}
    last_col_idx = ws.max_column
    last_col_letter = get_column_letter(last_col_idx)

    # Скрываем колонки (style и otrasl_ord)
    #ws.column_dimensions[last_col_letter].hidden = True
    for col_name, col_idx in headers.items():
        if col_name == 'otrasl_ord' or col_name == 'style':
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].hidden = True

    # Сетка для данных (с 4-й строки)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws[f"A4:{last_col_letter}{ws.max_row}"]:
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Применяем стили по правилам
    years = [str(y) for y in range(yearfrom, yearto + 1)]
    agg_dict = {f'y{y}': 'sum' for y in years}
    agg_dict['prirost'] = 'sum'

    for row_idx in range(4, ws.max_row + 1):
        last_cell_value = ws.cell(row=row_idx, column=last_col_idx).value

        if last_cell_value in STYLES:
            for col_name in headers:
                col_idx = headers[col_name]
                cell = ws.cell(row=row_idx, column=col_idx)
                for attr, value in STYLES[last_cell_value].items():
                    setattr(cell, attr, value)

        for rule in RULES:
            if not rule["condition"](last_cell_value):
                continue

            # Определяем столбцы для применения стиля
            columns_to_apply = []

            if rule["columns"] == ["*"]:
                # Применяем ко всем столбцам, кроме исключенных
                exclude_columns = rule.get("exclude", [])
                columns_to_apply = [col for col in headers if col not in exclude_columns]
            else:
                # Применяем к указанным столбцам
                columns_to_apply = [col for col in rule["columns"] if col in headers]

            for col_name in columns_to_apply:
                col_idx = headers[col_name]
                cell = ws.cell(row=row_idx, column=col_idx)

                # Применяем стиль
                for attr, value in STYLES[rule["style"]].items():
                    setattr(cell, attr, value)

                # Для числовых ячеек применяем специальное форматирование
                if col_name in [f"y{y}" for y in range(yearfrom, yearto + 1)] + ["prirost"]:
                    if isinstance(cell.value, (float, int)):
                        cell.number_format = '#,##0.0;-#,##0.0;"-"'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
    # Для s02 ервое слово должно быть красным
    for row_idx in range(4, ws.max_row + 1):
        style_cell = ws.cell(row=row_idx, column=last_col_idx)
        if style_cell.value == "s02":
            contragent_cell = ws.cell(row=row_idx, column=headers['contragent'])
            contragent_value = contragent_cell.value

            if contragent_value and isinstance(contragent_value, str):
                # Разделяем текст на слова
                words = contragent_value.split()
                if words:
                    # Используем InlineFont для rich text
                    from openpyxl.cell.rich_text import TextBlock, CellRichText
                    from openpyxl.cell.text import InlineFont

                    first_word_font = InlineFont(rFont='Times New Roman', sz='20', b=True, color='FF0000')
                    rest_font = InlineFont(rFont='Times New Roman', sz='20', b=True, color='000000')

                    first_word = TextBlock(first_word_font, words[0])
                    rest_of_text = TextBlock(rest_font, ' ' + ' '.join(words[1:]) if len(words) > 1 else '')

                    contragent_cell.value = CellRichText([first_word, rest_of_text])

    # Переименование заголовков
    header_map = {
        'contragent': "Потребитель",
        'otrasl': "Отрасль",
        'dogovor': "Договор поставки газа",
        'tu': "ТУ",
        'ver_real': "Вероятность реализации проекта",
        'grpost': "Группа поставщиков",
        'infr': "По программе газификации",
        'stpotr': "Статус потребителя",
        'stgaz': "Начало отбора",
        'prirost': f"Прирост\n{yearfrom}-{yearto}"
    }
    column_settings = {
        'contragent': {'width': 136.38},
        'otrasl': {'width': 28},
        'ver_real': {'width': 40},
        'dogovor': {'width': 18},
        'tu': {'width': 16},
        'prirost': {'width': 22}
    }

    # Форматирование заголовка таблицы (4-я строка)
    header_font = Font(name='Times New Roman', size=20, bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for cell in ws[4]:
        col_letter = cell.column_letter
        if cell.value in column_settings:
            ws.column_dimensions[col_letter].width = column_settings[cell.value]['width']
        elif cell.value and cell.value.startswith('y'):
            ws.column_dimensions[col_letter].width = 21

        if cell.value in header_map:
            cell.value = header_map[cell.value]
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        elif cell.value and cell.value.startswith("y"):
            cell.value = f"{cell.value[1:]} год"
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Применяем стиль к заголовку таблицы
        cell.font = header_font
        cell.alignment = header_alignment

    # Устанавливаем высоту для строки с заголовком таблицы
    ws.row_dimensions[4].height = 77

    # Включаем автоподбор высоты для всех строк
    for row in range(5, ws.max_row + 1):
        ws.row_dimensions[row].height = None  # Автоподбор высоты

    # Принудительно пересчитываем высоту строк
    for row in ws.iter_rows(min_row=5):
        max_height = 30  # Минимальная высота
        for cell in row:
            if cell.value and isinstance(cell.value, str) and '\n' in cell.value:
                line_count = cell.value.count('\n') + 1
                cell_height = line_count * 15  # Примерная высота на строку
                max_height = max(max_height, cell_height)

        if max_height > 30:
            ws.row_dimensions[cell.row].height = max_height

    wb.save(basic_path + name)
    return name, basic_path