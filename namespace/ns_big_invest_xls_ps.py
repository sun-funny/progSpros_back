import openpyxl
import io
import pandas as pd
import json
from openpyxl import Workbook
from flask import jsonify, session, request, send_from_directory, send_file
from progSpros_back.database_ps import db
from flask_restx import Namespace, Resource
from openpyxl import Workbook, load_workbook
from sqlalchemy import column

# Import the database session
from progSpros_back.database_ps import cache, errorhandler
from progSpros_back.functions.chart_data_functions_ps import apply_dynamic_filters
from progSpros_back.functions.query_functions_ps import big_invest_query_potr, query_prirost_potr_table
from progSpros_back.functions.utility_functions_ps import create_filter_params, substitute_in_json, sum_prirost, \
    set_db_connection, mapping
from progSpros_back.model.db_models_ps import Prirost, reference_models, Otrasl, FedState, Regions, GroupPost, \
    Contragent, StPotr, StGaz, PG, Dogovor, TU, Infr, VersProgn
from progSpros_back.model.mappings_ps import yn_mapping

# Define the namespace
ns_big_invest_xls_ps = Namespace('BigInvestXls', description='Крупные инвестиционные проекты в Excel')


def get_xlsx_from_db(download_name):
    pass


@ns_big_invest_xls_ps.route('/big_invest_xls_ps')
@ns_big_invest_xls_ps.response(200, 'Success')
@ns_big_invest_xls_ps.doc(params={
    'yearfrom': {'description': 'Год с', 'in': 'query', 'type': 'integer'},
    'yearto': {'description': 'Год по', 'in': 'query', 'type': 'integer'},
    'sum_pr': {'description': 'Прирост с', 'in': 'query', 'type': 'integer'},
    'otrasl': {'description': 'Отрасль', 'in': 'query', 'type': 'string'},
    'vers': {'description': 'Версия прогноза', 'in': 'query', 'type': 'string'},
    'grpost': {'description': 'Группа поставщиков', 'in': 'query', 'type': 'string'},
    'fo': {'description': 'Федеральный округ', 'in': 'query', 'type': 'string'},
    'region': {'description': 'Регион', 'in': 'query', 'type': 'string'},
    'dogovor': {'description': 'Договор', 'in': 'query', 'type': 'string'},
    'tu': {'description': 'ТУ', 'in': 'query', 'type': 'string'},
    'infr': {'description': 'Инфраструктура', 'in': 'query', 'type': 'string'}
})

class BigInvestXls(Resource):
    def get(self):
        """
        Возвращает обратно данные для карты

        Аргументы:
            - принимает аргумент global_filters из /get_basic_filters
            - принимает аргументы yearfrom, yearto
        """
        try:
            #db = set_db_connection()
            # Получить фильтр-параметры из запроса
            filter_params = create_filter_params(request)
            # Если не заданы глобальные параметры, взять их из session
            if not filter_params:
                filter_params = session.get('filter_params')

            # Мэппинги из справочников
            otr_mapping = mapping(Otrasl)
            vers_mapping = mapping(VersProgn)
            grpost_mapping = mapping(GroupPost)
            fo_mapping = mapping(FedState)
            region_mapping = mapping(Regions)

            # Определите базовый запрос с помощью динамических фильтров
            base_query = db.query(Prirost)
            base_query = apply_dynamic_filters(base_query, Prirost, filter_params, db, reference_models)

            yearfrom = request.args.get('yearfrom', 2023, type=int)
            yearto = request.args.get('yearto', 2034, type=int)
            sum_pr = request.args.get('sum_pr', -10000, type=int)

            # Параметры Отрасль
            reverse_otr_mapping = {value: key for key, value in otr_mapping.items()}
            otr = [company.strip() for item in request.args.getlist('otrasl', None) for company in item.split(',')]
            mapped_otr = [reverse_otr_mapping.get(company, company) for company in otr]
            if otr:
                base_query = base_query.filter((Prirost.tab_otrasl_economy_d314_ids.in_(mapped_otr)))

            # Параметры Версия прогноза
            reverse_vers_mapping = {value: key for key, value in vers_mapping.items()}
            vers = [company.strip() for item in request.args.getlist('vers', None) for company in item.split(',')]
            mapped_vers = [reverse_vers_mapping.get(company, company) for company in vers]
            if vers:
                base_query = base_query.filter((Prirost.tab_ver_real_pr_d314_ids.in_(mapped_vers)))

            # Параметры Группа поставщиков
            reverse_grpost_mapping = {value: key for key, value in grpost_mapping.items()}
            grpost = [company.strip() for item in request.args.getlist('grpost', None) for company in item.split(',')]
            mapped_grpost = [reverse_grpost_mapping.get(company, company) for company in grpost]
            if grpost:
                base_query = base_query.filter((Prirost.tab_group_post_d314_ids.in_(mapped_grpost)))

            # Параметры Федеральный округ
            reverse_fo_mapping = {value: key for key, value in fo_mapping.items()}
            fo = [company.strip() for item in request.args.getlist('fo', None) for company in item.split(',')]
            mapped_fo = [reverse_fo_mapping.get(company, company) for company in fo]
            if fo:
                base_query = base_query.filter((Prirost.tab_fo_d314_ids.in_(mapped_fo)))

            # Параметры Регион
            reverse_region_mapping = {value: key for key, value in region_mapping.items()}
            region = [company.strip() for item in request.args.getlist('region', None) for company in item.split(',')]
            mapped_region = [reverse_region_mapping.get(company, company) for company in region]
            if region:
                base_query = base_query.filter((Prirost.tab_region_d314_ids.in_(mapped_region)))

            # Параметры Договор
            reverse_dog_mapping = {value: key for key, value in yn_mapping.items()}
            dogovor = [company.strip() for item in request.args.getlist('dogovor', None) for company in item.split(',')]
            mapped_dog = [reverse_dog_mapping.get(company, company) for company in dogovor]
            if dogovor:
                base_query = base_query.filter((Prirost.tab_dogovor_visual_d314_ids.in_(mapped_dog)))

            # Параметры ТУ
            reverse_tu_mapping = {value: key for key, value in yn_mapping.items()}
            tu = [company.strip() for item in request.args.getlist('tu', None) for company in item.split(',')]
            mapped_tu = [reverse_tu_mapping.get(company, company) for company in tu]
            if tu:
                base_query = base_query.filter((Prirost.tab_tu_visual_d314_ids.in_(mapped_tu)))

            # Параметры Инфраструктура
            reverse_infr_mapping = {value: key for key, value in yn_mapping.items()}
            infr = [company.strip() for item in request.args.getlist('infr', None) for company in item.split(',')]
            mapped_infr = [reverse_infr_mapping.get(company, company) for company in infr]
            if infr:
                base_query = base_query.filter((Prirost.tab_infr_d314_ids.in_(mapped_infr)))

            # Продолжить создавать основной запрос
            query = big_invest_query_potr(base_query, Prirost, Otrasl, FedState, Regions, GroupPost, StPotr, StGaz, Infr,
                                     Dogovor, TU, yearfrom, yearto, Contragent)

            title = f"Крупные инвестиционные проекты"

            # Создать структуру вывода для Json
            result = []

            for row in query:

                if row.dogovor == 'V':
                    dog = '+'
                else:
                    dog = '-'

                if row.infr == 'V':
                    infr = '+'
                else:
                    infr = '-'

                if row.tu == 'V':
                    tu = '+'
                else:
                    tu = '-'

                if float(row.prirost) >= sum_pr  and float(row.prirost) != 0:
                    result_dict = {'Потребитель': row.contragent,
                                   'Федеральный округ': row.fo,
                                   'Регион': row.region,
                                   'Группа поставщиков': row.grpost,
                                   'Отрасль': row.otrasl,
                                   'Статус': row.stpotr,
                                   'Начало отбора': row.stgaz,
                                   'ПГ': infr,
                                   'Договор': dog,
                                   'ТУ': tu,
                                   'Прирост, млн м3': float(row.prirost)
                               }

                    result.append(result_dict)
            def get_prirost(element):
                return element['Прирост, млн м3']

            result.sort(key=get_prirost, reverse=True)
            y = json.dumps(result)
            df = pd.read_json(y)
            df.to_excel('/opt/foresight/progSpros_back/output.xlsx', index=False)

            # Обработка Excel
            workbook = openpyxl.load_workbook("/opt/foresight/progSpros_back/output.xlsx")
            sheet = workbook["Sheet1"]  # Получение листа по имени
            sheet.column_dimensions['A'].width = 40
            sheet.column_dimensions['B'].width = 37
            sheet.column_dimensions['C'].width = 31
            sheet.column_dimensions['D'].width = 20
            sheet.column_dimensions['E'].width = 20
            sheet.column_dimensions['F'].width = 18
            sheet.column_dimensions['G'].width = 14
            sheet.column_dimensions['H'].width = 5
            sheet.column_dimensions['I'].width = 8
            sheet.column_dimensions['J'].width = 5
            sheet.column_dimensions['K'].width = 15
            for row in sheet.iter_rows(min_row=2, max_row=3000, min_col=11, max_col=11):
                for cell in row:
                    cell.number_format = '#,##0.0'

            workbook.save("/opt/foresight/progSpros_back/output.xlsx")

            # Передать файл Excel выгруженный на сервер
            name = 'output.xlsx'
            basic_path = '/opt/foresight/progSpros_back/'
            #return send_from_directory(directory=f"{basic_path}", path=name, as_attachment=True)

            response = send_from_directory(directory=f"{basic_path}", path=name, as_attachment=True)
            response.headers.add('Access-Control-Allow-Origin', '*')

            return response
        except Exception as e:
            ns_big_invest_xls_ps.abort(*errorhandler(e))